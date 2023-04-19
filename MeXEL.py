# 필요한 모듈들을 불러옵니다. 기본적인 UI 구성요소를 제공하는 위젯 (클래스)들은 PyQt5.QtWidgets 모듈에 포함되어 있습니다.
# 폴더 내 파일 읽기 시 필요: os
import sys, os, fnmatch, requests, zipfile, tempfile, subprocess
import pandas as pd
from packaging import version
from datetime import datetime
from PyQt5.QtWidgets import QGroupBox, QMessageBox, QApplication, QMainWindow, QWidget, QPushButton, QVBoxLayout, QFileDialog, QToolTip, QHBoxLayout, QVBoxLayout, QLineEdit, QLabel, QSizePolicy, QComboBox
from PyQt5.QtGui import QFont, QIcon
from PyQt5.QtCore import Qt, QTimer, pyqtSlot
from PyQt5 import QtWidgets, QtCore
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# 앱의 현재 버전 정보
CURRENT_VERSION = "v1.1.1"

# 원격 서버의 API 주소
#API_URL = "https://example.com/api/check_update"

# GitHub 레포지토리 정보
GITHUB_API_URL = "https://api.github.com/repos/ByeongsuKim/MeXEL/releases/latest"


def check_and_update():
    latest_version, download_url = check_update()
    if latest_version:
        reply = QMessageBox.question(None, "업데이트 확인",
                                     f"새로운 버전 {latest_version}이(가) 발견되었습니다. 업데이트를 진행하시겠습니까?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

        if reply == QMessageBox.Yes:
            download_and_install_update(latest_version, download_url)
    else:
        print("최신 버전을 사용 중입니다.")      

def check_update():
    try:
        response = requests.get(GITHUB_API_URL)
        response.raise_for_status()
        data = response.json()
        latest_version = data["tag_name"]        
        if version.parse(latest_version) > version.parse(CURRENT_VERSION):           
            print("lasted version ", latest_version)
            print("current version ", CURRENT_VERSION)            
            # 올바른 ZIP 파일을 가리키는 다운로드 URL을 반환
            download_url = None
            for asset in data["assets"]:
                if asset["name"].endswith(".zip"):
                    download_url = asset["browser_download_url"]
                    break            
            if download_url:
                return latest_version, download_url
            else:
                raise Exception("No ZIP file found in the release assets.")                                    
            #return latest_version, data["assets"][0]["browser_download_url"]
    except requests.exceptions.RequestException:
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Warning)        
        msgBox.setWindowTitle("실패")
        msgBox.setText("업데이트 확인에 실패했습니다.")
        msgBox.exec_()
    except Exception as e:
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Warning)        
        msgBox.setWindowTitle("실패")
        msgBox.setText(str(e))
        msgBox.exec_()
    return None, None


def download_and_install_update(latest_version, download_url):
    try:
        response = requests.get(download_url, stream=True)
        response.raise_for_status()
        with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as tmp_file:
            for chunk in response.iter_content(chunk_size=8192):
                tmp_file.write(chunk)

        # 다운로드한 파일이 실제 ZIP 파일인지 확인하세요.
        with open(tmp_file.name, "rb") as f:
            file_signature = f.read(4)
        if file_signature != b'\x50\x4b\x03\x04':  # ZIP 파일의 시그니처 (PK\03\04)와 비교
            raise zipfile.BadZipFile("File is not a zip file")

        with zipfile.ZipFile(tmp_file.name, "r") as zip_ref:
            zip_ref.extractall(os.path.dirname(sys.executable))
        os.unlink(tmp_file.name)
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)        
        msgBox.setWindowTitle("성공")
        msgBox.setText(f"새 버전 {latest_version}이 설치되었습니다. 앱을 재시작해주세요.")
        msgBox.exec_()
    except requests.exceptions.RequestException:
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Warning)        
        msgBox.setWindowTitle("실패")
        msgBox.setText("업데이트 다운로드에 실패했습니다.")
        msgBox.exec_()

# 로고 표시를 위한 함수
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

class MyApp(QMainWindow):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        check_and_update()

        #아이콘
        myIcon = resource_path('logo.ico')
        self.setWindowIcon(QIcon(myIcon))
        #self.setMaximumSize(self.width(), self.height())
        self.setWindowTitle(f'맥셀(MErge eXEL) {CURRENT_VERSION}')  
        #최대화를 못하게 설정
        self.setWindowFlags(Qt.WindowCloseButtonHint | Qt.WindowMinimizeButtonHint)
        #상태표시줄
        self.statusBar().showMessage('문의: code.ssu@gmail.com')
        # 7초 후에 '준비 상태'로 변경
        QTimer.singleShot(7000, lambda: self.statusBar().showMessage('준비 상태'))

        # 병합할 행
        rows = 10
        cols = 3
        #self.mergeInfo = [[0 for j in range(cols)] for i in range(rows)]
        self.mergeInfo = []
        self.btnRemoveLayout = []
        # UID는 Unique ID의 뜻으로 mergeInfo(병합할 영역 정보)에 함께 들어가는 고유Key값으로 사용
        self.UID = 0
        # 병합될 영역 보여주는 칸에서 순서를 의미함
        self.ORDER = 0

        self.btn_remove = []


        print("MergeInfo : ", self.mergeInfo)        

        #툴팁 폰트 설정
        QToolTip.setFont(QFont('malgun', 10))

        #폴더 선택 버튼
        btn_folder = QPushButton(self)
        btn_folder.setText('폴더 선택')
        btn_folder.setToolTip('병합할 엑셀 파일이 있는 폴더를 선택해주세요.')
        btn_folder.clicked.connect(self.fileopen)

        #텍스트 상자
        self.textbox = QLineEdit(self)
        self.textbox.setReadOnly(True)

        #-----레이아웃: 폴더 선택 버튼 행
        groupbox_folder = QGroupBox("병합 파일 폴더 선택")
        groupbox_folder.setStyleSheet("QGroupBox { border: 1px solid gray;}")
        groupbox_folder.setFixedHeight(100)
        hbox_folder = QHBoxLayout()
        hbox_folder.addWidget(btn_folder)
        hbox_folder.addWidget(self.textbox)
        hbox_folder.setSpacing(0)
        groupbox_folder.setLayout(hbox_folder)

        # 시작, 마지막 버튼
        row_layout = QHBoxLayout()
        
        sheet_label = QLabel('시트 ', self)
        self.sheet_input = QComboBox(self)
        for i in range(1, 51):
            self.sheet_input.addItem(str(i))
        self.sheet_input.setEditable(True)
        #self.sheet_input.currentTextChanged.connect(self.update_end_input)  # 시작 행 변경 시 이벤트 발생

        start_label = QLabel('영역 ', self)
        self.start_input = QComboBox(self)
        for i in range(1, 51):
            self.start_input.addItem(str(i))
        self.start_input.setEditable(True)
        self.start_input.currentTextChanged.connect(self.update_end_input)  # 시작 행 변경 시 이벤트 발생
        
        end_label = QLabel('~', self)
        self.end_input = QComboBox(self)
        self.end_input.addItem('데이터 끝')
        for i in range(1, 51):
            self.end_input.addItem(str(i))
        self.end_input.setEditable(True)
        self.row_label = QLabel('행', self)

        space_label = QLabel('       ')

        # 추가 버튼
        btn_add = QPushButton(self)
        btn_add.setText('병합 영역 추가')
        btn_add.setToolTip('해당 영역을 설정합니다.')
        btn_add.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn_add.clicked.connect(self.addArea)

        #-----레이아웃: 시작, 마지막, 추가 버튼 레이아웃
        groupbox_sheetrow = QGroupBox("영역 설정하기")
        groupbox_sheetrow.setStyleSheet("QGroupBox { border: 1px solid gray;}")
        groupbox_sheetrow.setFixedHeight(100)

        row_layout.addWidget(sheet_label)
        row_layout.addWidget(self.sheet_input)
        row_layout.addWidget(space_label)
        row_layout.addWidget(start_label)
        row_layout.addWidget(self.start_input)
        row_layout.addWidget(end_label)
        row_layout.addWidget(self.end_input)
        row_layout.addWidget(self.row_label)
        row_layout.addWidget(space_label)
        row_layout.addWidget(btn_add)


        #-----레이아웃: 선택된 병합 영역
        '''
        groupbox_selectedRow = QGroupBox("병합할 영역")
        groupbox_selectedRow.setStyleSheet("QGroupBox { border: 1px solid gray;}")
        self.selectedRow_layout = QVBoxLayout()
        self.selectedRow_layout.addSpacing(5)
        '''
        self.centralwidget = QtWidgets.QWidget(self)
        self.centralwidgetLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.scrollArea = QtWidgets.QScrollArea(self.centralwidget)
        self.scrollArea.setWidgetResizable(True)
        self.scrollAreaWidget = QtWidgets.QWidget()
        self.scrollAreaWidget.setGeometry(QtCore.QRect(0, 0, 600, 800))
        self.scrollAreaWidgetLayout = QtWidgets.QVBoxLayout(self.scrollAreaWidget)
        self.scrollAreaWidgetLayout.addItem(QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding))
        self.scrollArea.setWidget(self.scrollAreaWidget)
        self.scrollArea.setMinimumHeight(200)

        # -----레이아웃: 병합 버튼
        merge_layout = QVBoxLayout()
        merge_layout.addLayout(row_layout)
        groupbox_sheetrow.setLayout(merge_layout)
        groupbox_sheetrow.setStyleSheet("QGroupBox { border: 1px solid gray;}")
        
        # 병합 실행 버튼
        btn_run = QPushButton(self)
        btn_run.setText('병합 실시')
        btn_run.setToolTip('병합을 실시합니다.')
        btn_run.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        btn_run.clicked.connect(self.mergeExcel)

        # -----레이아웃: Main        
        mainWidget = QWidget(self)
        self.setCentralWidget(mainWidget)
        main_layout = QVBoxLayout()
        main_layout.addWidget(groupbox_folder)
        main_layout.addWidget(groupbox_sheetrow)
        main_layout.addWidget(self.scrollArea)
        main_layout.addWidget(btn_run)
        mainWidget.setLayout(main_layout)

        self.setGeometry(300, 300, 350, 400)
        self.show()


    def check_version():
        curr



    def addArea(self):
        if len(self.mergeInfo) > 19:
            self.statusBar().showMessage('최대 20개의 영역만 추가 가능합니다.')
            return
        else:
            #텍스트 필드 값 가져오기
            sheet_no = int(self.sheet_input.currentText())
            start_row = int(self.start_input.currentText())
            if(self.end_input.currentText()=='데이터 끝'):
                end_row = self.end_input.currentText()
            else:
                end_row = int(self.end_input.currentText())
            
            # 현재 self.ORDER 순서를 1 증가시키고 self.mergeInfo에 추가
            # 즉, self.UID는 버튼과의 1:1 매칭, self.ORDER는 현재 화면상 그룹박스 순서
            
            self.ORDER = self.ORDER + 1
            self.mergeInfo.append([self.UID, self.ORDER, sheet_no, start_row, end_row])            


            #그룹박스 추가하기
            #그룹박스 - 병합 시트, 행 정보

            count = self.scrollAreaWidgetLayout.count() - 1
            groupBox = QtWidgets.QGroupBox(self.scrollAreaWidget)
            self.scrollAreaWidgetLayout.insertWidget(count, groupBox)

            #그룹박스 - 삭제 버튼
            print("UID : ", self.UID)
            btn =  QPushButton(self)
            btn.setText('제거')
            btn.id = self.UID
            btn.setToolTip('설정한 영역을 삭제합니다.')
            btn.setFixedWidth(150)
            btn.clicked.connect(lambda: self.deleteLaterGroupBox(btn.id))
            #self.btn_remove[self.UID].append([btn])
            '''
            self.btn_remove[self.UID].setText('제거')
            self.btn_remove[self.UID].id = self.UID
            self.btn_remove[self.UID].setToolTip('설정한 영역을 삭제합니다.')
            self.btn_remove[self.UID].setFixedWidth(150)
            self.btn_remove[self.UID].clicked.connect(lambda: self.deleteLaterGroupBox(self.btn_remove[self.UID].id))
            '''
            
            #comboBox = QtWidgets.QComboBox(groupBox)
            #comboBox.addItems(['val1', 'val2', 'val3'])
            gridLayout = QtWidgets.QGridLayout(groupBox)
            info = "  • 시트 [" + str(sheet_no) + "]의    [" + str(start_row) + "] ~ [" + str(end_row) + "]행"

            gridLayout.addWidget(QtWidgets.QLabel(info, groupBox),       0, 0, 1, 1)
            gridLayout.addWidget(btn, 0, 1, 1, 1)
            #gridLayout.addWidget(comboBox,                                                1, 0, 1, 1)
            #gridLayout.addWidget(QtWidgets.QSlider(QtCore.Qt.Horizontal, groupBox),       1, 1, 1, 1)

            self.UID = self.UID + 1      
            
    def fileopen(self):
        global pathDir
        pathDir = QFileDialog.getExistingDirectory()
        self.textbox.setText(pathDir)
        extensions = ('*.xlsx', '*.xlsm', '*.xlsb', '*.csv')
        global fileList
        fileList = []
        if pathDir:
            fileList = [os.path.join(pathDir, file) for file in os.listdir(pathDir) if any(fnmatch.fnmatch(file, ext) for ext in extensions)]
            print(fileList)
            self.statusBar().showMessage(f'{len(fileList)}개의 엑셀 파일을 찾았습니다.')    
        else:
            self.statusBar().showMessage('폴더를 선택해주세요.')

    def deleteLaterGroupBox(self, index):
        '''
        [Example]
        Dynamically add and remove QWidgets to QLayout with PyQt5
        https://ymt-lab.com/en/post/2021/pyqt5-delete-widget-test/

        '''
        print(index)
        print(self.mergeInfo)
        #순서를 의미하는 것이므로.   
        targetNo = -1
        for i in range(len(self.mergeInfo)):
            if self.mergeInfo[i][0] == index:
                targetNo = i
                break
        #targetNo가 실제 현재 있는 병합영역의 순서를 의미함 0, 1, 2, 3 순
        self.mergeInfo.remove(self.mergeInfo[targetNo])
        for i in range(len(self.mergeInfo)):
            self.mergeInfo[i][1] = i
        self.ORDER = self.ORDER-1
        print("reMergeInfo :", self.mergeInfo)
        item = self.scrollAreaWidgetLayout.itemAt(targetNo)        
        widget = item.widget()
        widget.deleteLater()

    def mergeExcel(self):       
        if not self.textbox.text(): # 텍스트박스에 값이 없는 경우
            self.statusBar().showMessage('폴더를 먼저 선택해주세요.')
            return            
        if len(self.mergeInfo)==0:
            self.statusBar().showMessage('병합 영역을 추가해주세요.')
            return
        #기존 파일을 여는 wb
        wb = Workbook()
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f'merged_{timestamp}.xlsx'
        new_file = os.path.join(os.path.dirname(pathDir), filename)

        newWb = Workbook()

        # mergeInfo를 for문 돌며 같은 파일을 여러번 열어야 함.
        for i, extr in enumerate(self.mergeInfo):
            #사용한 데이터프레임
            df = pd.DataFrame()
            #시브번호, 병합행 시작, 끝 추출
            sheetno = extr[2]-1
            srow = extr[3]
            #
            #ws = wb.worksheets[sheetno]
            #merged_data에 각 파일의 해당 시트, 행을 누적함
            merged_data = []
            for file in fileList:
                #print(i, " - ", extr)
                #print(file)
                wb = load_workbook(file)
                try:
                    ws = wb.worksheets[sheetno]
                except Exception as e:
                    msgBox = QMessageBox()
                    msgBox.setIcon(QMessageBox.Warning)
                    index_slash = file.index("\\")
                    shortFname = file[index_slash+1:]
                    msgBox.setWindowTitle("시트 병합 오류")
                    msgBox.setText(shortFname + "에 " + str(sheetno+1) +"번 시트가 없어 병합을 중지합니다.")
                    msgBox.exec_()
                    # 7초 후에 '준비 상태'로 변경
                    QTimer.singleShot(5000, lambda: self.statusBar().showMessage('준비 상태'))
                    self.statusBar().showMessage(f'시트 읽기 중 오류:  {e}')
                    return
                if(str(extr[4])=='데이터 끝'):
                    erow = ws.max_row
                    if (erow<srow):
                        erow=srow
                else:
                    erow = extr[4]


                for row in range(srow, erow+1):
                    row_data = [ws.cell(row, col).value for col in range(1, ws.max_column+1)]
                    print('\t'.join(str(cell) for cell in row_data))
                    merged_data.append(row_data)
                    #여기에러df.append(pd.Series(row_data, index=df.columns), ignore_index=True)
                    # 참고: https://emilkwak.github.io/dataframe-list-row-append-ignore-index
                    #다른 방법으로 해결가능: 엑셀 열고 각 행 해당 시트에 계속 누적.
            #모든 루프가 다 돌면 merged_data에 모든 엑셀 파일의 특정 시트와 행 누적된 상태
            print("-------------------------")    
            print(merged_data)          
            newDf = pd.DataFrame(merged_data)
            newSht = newWb.create_sheet("Sheet"+str(sheetno+1)+"_"+str(extr[0]))
            for r in dataframe_to_rows(newDf, index=False, header=False):
                newSht.append(r)

        del newWb['Sheet']
        newWb.save(new_file)
        
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setWindowTitle("성공")
        msgBox.setText("병합이 완료되었습니다.\n병합된 결과를 확인하세요.")
        msgBox.exec_()
        os.startfile(new_file)


    def update_end_input(self):
        start_row = int(self.start_input.currentText())
        self.end_input.clear()
        self.end_input.addItem('데이터 끝')
        for i in range(start_row, 51):
            self.end_input.addItem(str(i))


if __name__ == '__main__':
   app = QApplication(sys.argv)
   ex = MyApp()
   ex.setGeometry(100, 100, 500, 400)
   sys.exit(app.exec_())

#
# 실행파일 만들기
# pyinstaller -w -F -i="logo.ico" --add-data="logo.ico;./" -n="MeXEL.exe" MeXEL.py